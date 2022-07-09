import speech_recognition as sr #speech recognizer module
import sounddevice as sd #sound producer module
import wavio #read and write function module
import os #module to interact with operating system
from tkinter import * #to make graphical user interface
from tkinter import messagebox
from PIL import Image, ImageTk #for manipulation of images
import time #module for accessing time
import pandas as pd  # data manipulation and analysis.
import datetime  # to get currrent date
from plyer import notification
import pyttsx3  # to speak written
import openpyxl,xlrd #module to access excel sheets
from openpyxl import Workbook
import pathlib
c=0
print("\t1.set remainder by typing\n\t2.set remainder by speech\n\t3.add an event to calender\n\t4.check for events planned\n")
c=int(input())
if c==1:
    p = Tk()
    p.title(' Desktop Notifier')
    p.geometry("500x300")
    #p.config(bg='#E3AEAE')
    img =Image.open("notify-label (1).png")
    tkimage = ImageTk.PhotoImage(img)


    # get details
    def get_info():
        get_head = title.get()
        get_message = msg.get()
        get_time = time1.get()
        # print(get_head,get_message, tt)

        if get_head == "" or get_message == "" or get_time == "":
            messagebox.showerror("Warning!!", "All fields are required!")
        else:
            int_time =float(get_time)
            min_to_sec = int_time * 60
            messagebox.showinfo("notifier set", "set notification ?")
            p.destroy()
            time.sleep(min_to_sec)

            notification.notify(title=get_head,
                                message=get_message,
                                app_name="Desktop Notifier",
                                app_icon="ico5.ico",
                                toast=True,
                                timeout=8)
            speak("USER!!"+get_message)


    engine = pyttsx3.init('sapi5')
    voices = engine.getProperty('voices')
    engine.setProperty('voice', voices[1].id)


    def speak(audio):
        engine.say(audio)
        engine.runAndWait()


    img_label = Label(p, image=tkimage).grid()

    # Label - Title
    p_label = Label(p, text="Title to Notification", font=("poppins", 10))
    p_label.place(x=12, y=70)

    # ENTRY - Title
    title = Entry(p, width="25", font=("poppins", 13))
    title.place(x=123, y=70)

    # Label - Message
    msg_label = Label(p, text="Display Message", font=("poppins", 10))
    msg_label.place(x=12, y=120)

    # ENTRY - Message
    msg = Entry(p, width="40", font=("poppins", 13))
    msg.place(x=123, height=30, y=120)

    # Label - Time
    time_label = Label(p, text="Set Time", font=("poppins", 10))
    time_label.place(x=12, y=175)

    # ENTRY - Time
    time1 = Entry(p, width="5", font=("poppins", 13))
    time1.place(x=123, y=175)

    # Label - min
    time_min_label = Label(p, text="min", font=("poppins", 10))
    time_min_label.place(x=175, y=180)

    # Button
    press= Button(p, text="SET NOTIFICATION", font=("poppins", 10, "bold"), fg="#ffffff", bg="#528DFF", width=20,
                 relief="raised",
                 command=get_info)
    press.place(x=170, y=230)

    p.resizable(0, 0)
    p.mainloop()
elif c==2:
    def voicerec():
        l = ""
        fps = 44100
        inetrval = 3
        print("Recognizing.....")

        recording = sd.rec(inetrval * fps, samplerate=fps, channels=2)
        sd.wait()
        print("Done")

        wavio.write('output.wav', recording, fps, sampwidth=2)
        rec = sr.Recognizer()
        audioF = 'output.wav'
        with sr.AudioFile(audioF) as sourceF:
            audio = rec.record(sourceF)
            print("File is READING")
        print("file text is : ")
        try:
            text = rec.recognize_google(audio)
            print(text)
            for i in text:
                if i != ' ':
                    l = l + i
                else:
                    break
            return float(l)



        except Exception as e:
            print(e)

        os.remove('output.wav')


    def voicerec1():
        l = ""
        fps = 44100
        inetrval = 5
        print("Recognizing speak.....")

        recording = sd.rec(inetrval * fps, samplerate=fps, channels=2)
        sd.wait()
        print("Done")

        wavio.write('output.wav', recording, fps, sampwidth=2)
        rec = sr.Recognizer()
        audioF = 'output.wav'
        with sr.AudioFile(audioF) as sourceF:
            audio = rec.record(sourceF)
            print("File is READING")
        print("file text is : ")
        try:
            text = rec.recognize_google(audio)
            print(text)
            return text



        except Exception as e:
            print(e)

        os.remove('output.wav')


    p = Tk()
    p.title('Desktop Notifier')
    p.geometry("500x300")
    img = Image.open("notify-label (1).png")
    tkimage = ImageTk.PhotoImage(img)


    # get details
    def get_details():
        engine = pyttsx3.init('sapi5')
        voices = engine.getProperty('voices')
        engine.setProperty('voice', voices[1].id)

        def speak(audio):
            engine.say(audio)
            engine.runAndWait()
        print("SPELL THE TITLE:")
        get_head = voicerec1()
        print("\n\n")
        print("SPELL THE MESSAGE:")
        get_message = voicerec1()
        print("\n\n")
        print("SPELL THE TIME IN MINUTES:")
        get_time = voicerec()
        print("\n\n")

        if get_head == "" or get_message == "" or get_time == "":
            messagebox.showerror("Warning!!", "All fields are required!")
        else:
            int_time = float(get_time)
            min_to_sec = int_time * 60
            messagebox.showinfo("notifier set", "set notification ?")
            p.destroy()
            time.sleep(min_to_sec)

            notification.notify(title=get_head,
                                message=get_message,
                                app_name="Notifier",
                                app_icon="ico5.ico",
                                toast=True,
                                timeout=8)
            speak("USER!!" +str(get_message))




    img_label = Label(p, image=tkimage).grid()

    msg_label = Label(p, text="Setting Remainder through speech...\nclick SET NOTIFICATION button and start the speech!!", font=("poppins", 12,"bold"))
    msg_label.place(x=60, y=120)

    # Button
    but = Button(p, text="SET NOTIFICATION", font=("poppins", 10, "bold"), fg="#ffffff", bg="#528DFF", width=20,
                 relief="raised",
                 command=get_details)
    but.place(x=170, y=230)

    p.resizable(0, 0)
    p.mainloop()
elif c==3:
    p = Tk()
    p.title("Desktop Event Remainder")
    p.geometry("500x300")
    img = Image.open("notify-label (1).png")
    tkimage = ImageTk.PhotoImage(img)
    file = pathlib.Path("dates.xlsx")
    if file.exists():
        pass
    else:
        file = Workbook()
        sheet = file.active
        sheet["A1"] = "name"
        sheet["B1"] = "date"
        file.save("dates.xlsx")


    def set_not():
        y = name.get()
        z = date.get()
        print(y)
        print(z)
        print("Remainder set successfully!!")
        file = openpyxl.load_workbook("dates.xlsx")
        sheet = file.active
        sheet.cell(column=1, row=sheet.max_row + 1, value=y)
        sheet.cell(column=2, row=sheet.max_row, value=z)
        file.save("dates.xlsx")


    img_label = Label(p, image=tkimage).grid()
    Label(p, text="Name of Event:",font=("poppins", 10)).place(x=12, y=70)
    Label(p, text="Date of Event:",font=("poppins", 10)).place(x=12, y=120)
    name = Entry(p, width="25", font=("poppins", 13))
    name.place(x=123, y=70)
    date = Entry(p, width="25", font=("poppins", 13))
    date.place(x=123, y=120)
    Button(p, text="SET NOTIFICATION", font=("poppins", 10, "bold"), fg="#ffffff", bg="#528DFF", width=20,
           relief="raised",
           command=set_not).place(x=170, y=230)
    p.mainloop()
if c==4:
    engine = pyttsx3.init('sapi5')
    voices = engine.getProperty('voices')
    engine.setProperty('voice', voices[1].id)


    def speak(audio):
        engine.say(audio)
        engine.runAndWait()


    def notification1(title, msg):
        notification.notify(
            title=title,
            message=msg,
            app_icon="2.ico",
            timeout=8
        )

    readf = pd.read_excel("dates.xlsx")
    today = datetime.datetime.now().strftime("%d-%m-%y")
    print("to day's date:",today)
    for index, item in readf.iterrows():
        da = item["date"]
        if today == da:
            a = item["name"]
            notification1("REMAINDER ALERT", a)
            speak("USER!!" + a)
            time.sleep(3)
